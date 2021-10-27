from datetime import datetime as dt
from openpyxl import load_workbook, utils as xl_utils
import pandas as pd
import re

from base64 import b64encode as enc64

from flask import jsonify
import json
from jsonschema import validate, exceptions as exc
from requests import auth



def select_subdict(a_dict, sub_keys):
    if not sub_keys.is_subset(a_dict.keys()):
        raise ValueError('Trying to select a subdict with keys not contained on large dict.')
    small_dict = dict((k, a_dict[k]) for k in sub_keys)
    return small_dict


def str_camel_to_snake(cameled:str):
    subbed = re.sub('(.)([A-Z][a-z]+)',  r'\1_\2', cameled)
    snaked = re.sub('([a-z0-9])([A-Z])', r'\1_\2', subbed).lower()
    return snaked


def str_snake_to_camel(snaked:str, first_word_too=False):
    splitted    = snaked.split("_")
    first_word  = splitted.pop(0)
    first_camel = first_word.title() if first_word_too else first_word.lower()
    cameled     = first_camel + "".join(word.title() for word in splitted)
    return cameled


def snake_2_camel(snake_str):
    first, *others = snake_str.split('_')
    return ''.join([first.lower(), *map(str.title, others)])


def camelize_dict(snake_dict):
    if not isinstance(snake_dict, dict):
        return snake_dict

    def pass_value(a_val): 
        if isinstance(a_val, list):
            passed = list(map(camelize_dict, a_val))
        elif isinstance(a_val, dict):
            passed = camelize_dict(a_val)
        else:
            passed = a_val 
        return passed

    new_dict = dict((k, pass_value(v)) for (k, v) in snake_dict.items())
    return new_dict


def shortcut_target(filename, file_ext=None):
    def ext_regex(file_ext):
        if file_ext is None: 
            file_ext = "xlsx"
        if isinstance(file_ext, str):
            ext_reg = file_ext
        elif isinstance(file_ext, list):
            ext_reg = f"{'|'.join(file_ext)}"
        else:
            raise "FILE_EXT format is not supported."
        return ext_reg
    
    file_regex = fr"C:\\.*\.{ ext_regex(file_ext) }"
    with open(filename, "r", encoding="ISO-8859-1") as _f: 
        a_path = re.findall(file_regex, _f.read(), flags=re.DOTALL)

    if len(a_path) != 1: 
        raise "Not unique or No shortcut targets found in link."
    return a_path[0]


def read_excel_table(file, sheet, table): 
    try:
        a_wb = load_workbook(file, data_only=True)
    except xl_utils.exceptions.InvalidFileException: 
        a_wb = load_workbook(shortcut_target(file), data_only=True)
    a_ws  = a_wb[sheet]
    a_tbl = a_ws.tables[table]
    
    rows_ls = [[ cell.value for cell in row ] for row in a_ws[a_tbl.ref]]
    tbl_df  = pd.DataFrame(data=rows_ls[1:], index=None, columns=rows_ls[0])
    return tbl_df



def dict_minus(dict, key, copy=True): 
    b_dict = dict.copy() if copy else dict
    b_dict.pop(key)
    return b_dict


def encode64(a_str): 
    encoded = enc64(a_str.encode('ascii')).decode('ascii')
    return encoded


class BearerAuth(auth.AuthBase):
    def __init__(self, token):
        self.token = token
    def __call__(self, req):
        req.headers["authorization"] = f"Bearer {self.token}"
        return req


def validate_input(payload, input_file):
    try: 
        with open(input_file, 'r') as _f:
            input_schema = json.load(_f)
        validate(instance=payload, schema=input_schema)
        an_object = {"error" : False}
    except (exc.ValidationError, exc.SchemaError) as err:
        response = {
            "code"          : "0001",
            "type"          : "validation/input",
            "status_code"   : "400",
            "timestamp"     : str(dt.now()),
            "instance"      : "input/messages_strategy/invalid_structure",
            "detail"        : str(err) }
        an_object = {"error" : 400, "output" : (jsonify(response), 400)}
    return an_object


def request_from_response(resp_request): 
    hdrs_items  = "\r\n".join(f"{k}:{v}" for (k, v) in resp_request.headers.items())
    request_str = f"{resp_request.method} {resp_request.url}\n{hdrs_items}\r\n\r\n{resp_request.body}"

    return request_str



    